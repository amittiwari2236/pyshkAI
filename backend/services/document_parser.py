import pdfplumber
import docx
import openpyxl

def extract_text_from_pdf(file_path):
    text = ""
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                extracted = page.extract_text()
                if extracted:
                    text += extracted + "\\n"
    except Exception as e:
        print(f"PDF Extraction Error: {e}")
    return text

def extract_text_from_docx(file_path):
    text = ""
    try:
        doc = docx.Document(file_path)
        for para in doc.paragraphs:
            text += para.text + "\\n"
    except Exception as e:
        print(f"Word Extraction Error: {e}")
    return text

def extract_text_from_excel(file_path):
    text = ""
    try:
        wb = openpyxl.load_workbook(file_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_str = " ".join([str(cell) for cell in row if cell is not None])
                if row_str.strip():
                    text += row_str + "\\n"
    except Exception as e:
        print(f"Excel Extraction Error: {e}")
    return text

def process_document(document):
    """
    Given a KnowledgeDocument instance, extract text based on document_type
    and save it to the extracted_text field.
    """
    if not document.file:
        return False
        
    file_path = document.file.path
    text = ""
    
    if document.document_type == 'PDF':
        text = extract_text_from_pdf(file_path)
    elif document.document_type == 'Word':
        text = extract_text_from_docx(file_path)
    elif document.document_type == 'Excel':
        text = extract_text_from_excel(file_path)
    elif document.document_type == 'Text':
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
        except:
            pass
            
    document.extracted_text = text
    document.save()
    return True
